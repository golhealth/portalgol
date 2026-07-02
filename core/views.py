from django.conf import settings
from django.contrib import messages
from django.contrib.auth.views import PasswordChangeView
from django.core.exceptions import PermissionDenied
from django.core.mail import send_mail
from django.contrib.auth import (
    authenticate,
    get_user_model,
    login,
    update_session_auth_hash,
)
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import PasswordChangeForm, SetPasswordForm
from django.db import transaction
from django.db.models import Count, Q, Value
from django.db.models.functions import Coalesce
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_http_methods, require_POST

from datetime import date, datetime, timedelta
from io import BytesIO
from urllib.parse import urlencode
import unicodedata
import openpyxl

from .forms import (
    CategoriaForm,
    CriarUtilizadorForm,
    EditarUtilizadorForm,
    DepartamentoForm,
    EpocaAtivoForm,
    EpocaForm,
    ImportArbitrosExcelForm,
    ModalidadeForm,
)
from .models import (
    Categoria,
    Departamento,
    Epoca,
    EpocaArbitro,
    EpocaAssociacao,
    Modalidade,
    PerfilConta,
)
from .password_utils import gerar_palavra_passe_segura, username_a_partir_do_email
from .site_url import absolute_url
from exames.forms import ExameForm
from associacao.models import Associacao
from arbitro.models import Arbitro, RealizacaoExameArbitro
from exames.models import MarcacaoExame
from prestador.models import Prestador

User = get_user_model()


@require_http_methods(["GET", "POST"])
@ensure_csrf_cookie
def login_view(request):
    """
    Página de login (página inicial da aplicação) com fluxo de 2 passos:
    1) E-mail + password -> gera código MFA
    2) Introdução do código MFA -> conclui login
    """
    # Já autenticado e não é timeout -> vai para dashboard
    if request.user.is_authenticated and "timeout" not in request.GET:
        return redirect("core:dashboard")

    session_expired = "timeout" in request.GET
    is_mfa_step = bool(request.session.get("mfa_pending")) and "mfa" in request.GET
    mfa_pending = request.session.get("mfa_pending") or {}
    mfa_email = mfa_pending.get("email", "")
    error = ""
    email_value = ""
    dev_mfa_code = None

    # Em desenvolvimento, podemos mostrar o código MFA no template
    if settings.DEBUG and mfa_pending.get("code"):
        dev_mfa_code = mfa_pending.get("code")

    # POST: ou valida MFA ou valida credenciais
    if request.method == "POST":
        # Passo 2 - validação do código MFA
        if request.session.get("mfa_pending") and "mfa_code" in request.POST:
            code = (request.POST.get("mfa_code") or "").strip()
            pending = request.session.get("mfa_pending") or {}
            expected_code = str(pending.get("code") or "")
            expires_at = int(pending.get("expires_at") or 0)
            now_ts = int(timezone.now().timestamp())

            if not expected_code or now_ts > expires_at:
                error = "O código de verificação expirou. Por favor, faça login novamente para receber um novo código."
                request.session.pop("mfa_pending", None)
                request.session.pop("mfa_attempts", None)
                is_mfa_step = False
            elif code != expected_code:
                attempts = int(request.session.get("mfa_attempts") or 0) + 1
                request.session["mfa_attempts"] = attempts
                if attempts >= 5:
                    error = "Foram introduzidos códigos incorretos em excesso. Faça login novamente para receber um novo código."
                    request.session.pop("mfa_pending", None)
                    request.session.pop("mfa_attempts", None)
                    is_mfa_step = False
                else:
                    error = "Código de verificação incorreto. Verifique o código enviado para o seu e-mail."
                    is_mfa_step = True
            else:
                # Código correto: concluir login
                request.session.pop("mfa_attempts", None)
                pending_user_id = pending.get("user_id")
                user = None
                if pending_user_id:
                    try:
                        user = User.objects.get(id=pending_user_id)
                    except User.DoesNotExist:
                        user = None
                    if user is not None:
                        login(request, user)
                request.session.pop("mfa_pending", None)
                if user is not None:
                    perfil, _ = PerfilConta.objects.get_or_create(
                        user=user,
                        defaults={"forcar_troca_password": False},
                    )
                    if perfil.forcar_troca_password:
                        return redirect("core:primeiro_acesso_password")
                return redirect("core:dashboard")

        # Passo 1 - e-mail + password
        else:
            email = (request.POST.get("email") or "").strip()
            password = request.POST.get("password") or ""
            email_value = email

            request.session.pop("mfa_pending", None)
            request.session.pop("mfa_attempts", None)

            user = None
            if email and password:
                try:
                    user_obj = User.objects.get(email__iexact=email)
                except User.DoesNotExist:
                    user_obj = None
                if user_obj is not None:
                    user = authenticate(request, username=user_obj.username, password=password)

            if not user:
                error = "E-mail ou palavra-passe incorretos."
            else:
                # Gerar código de 6 dígitos e guardar em sessão
                import random

                code = str(random.randint(0, 999999)).zfill(6)
                expires_at = int(timezone.now().timestamp()) + 600  # 10 minutos

                request.session["mfa_pending"] = {
                    "user_id": user.id,
                    "email": user.email or email,
                    "nome": getattr(user, "first_name", "") or user.get_username(),
                    "code": code,
                    "expires_at": expires_at,
                }
                request.session["mfa_attempts"] = 0

                subject = "Código de verificação - Gol Health"
                message = f"O seu código de verificação é: {code}"
                try:
                    send_mail(
                        subject,
                        message,
                        getattr(settings, "DEFAULT_FROM_EMAIL", None),
                        [request.session["mfa_pending"]["email"]],
                        fail_silently=False,
                    )
                except Exception:
                    if settings.DEBUG:
                        return redirect(f"{request.path}?mfa=1")
                    request.session.pop("mfa_pending", None)
                    request.session.pop("mfa_attempts", None)
                    error = (
                        "Nao foi possivel enviar o codigo MFA por e-mail. "
                        "Verifique a configuracao SMTP e tente novamente."
                    )
                else:
                    return redirect(f"{request.path}?mfa=1")

    context = {
        "session_expired": session_expired,
        "is_mfa_step": is_mfa_step,
        "mfa_email": mfa_email,
        "error": error,
        "email": email_value,
        "dev_mfa_code": dev_mfa_code,
    }
    return render(request, "core/login.html", context)


@login_required
def index_view(request):
    """Página inicial do backoffice (exige login)."""
    return render(request, "core/index.html")


def _arbitro_iniciais(nome: str) -> str:
    partes = [p for p in (nome or "").split() if p]
    if not partes:
        return "?"
    if len(partes) == 1:
        return (partes[0][:2] or "?").upper()
    return (partes[0][0] + partes[1][0]).upper()


def _badge_class_exame(nome: str) -> str:
    classes = (
        "bg-blue-100 dark:bg-blue-900/30 text-blue-700 dark:text-blue-300",
        "bg-green-100 dark:bg-green-900/30 text-green-700 dark:text-green-300",
        "bg-amber-100 dark:bg-amber-900/30 text-amber-700 dark:text-amber-300",
        "bg-purple-100 dark:bg-purple-900/30 text-purple-700 dark:text-purple-300",
        "bg-pink-100 dark:bg-pink-900/30 text-pink-700 dark:text-pink-300",
    )
    h = sum(ord(c) for c in (nome or "")) % len(classes)
    return classes[h]


@login_required
def dashboard_view(request):
    """
    Dashboard principal com contagens e listas a partir da base de dados.
    Certificação: ciclo anual a partir de data_ultimo_exame (limite = +365 dias).
    """
    epoca_atual = Epoca.objects.filter(ativo=True).order_by("-criado_em").first()
    today = timezone.now().date()

    total_arbitros = Arbitro.objects.count()
    prestadores_ativos = Prestador.objects.filter(estado="ativo").count()

    now_dt = timezone.now()
    start_mes = date(now_dt.year, now_dt.month, 1)
    if now_dt.month == 12:
        end_mes = date(now_dt.year, 12, 31)
    else:
        end_mes = date(now_dt.year, now_dt.month + 1, 1) - timedelta(days=1)
    exames_realizados_mes = RealizacaoExameArbitro.objects.filter(
        data_realizacao__gte=start_mes,
        data_realizacao__lte=end_mes,
    ).count()

    limite_vencido = today - timedelta(days=365)
    limite_avencer_max = today - timedelta(days=335)
    cert_vencidos = Arbitro.objects.filter(
        Q(data_ultimo_exame__isnull=True) | Q(data_ultimo_exame__lt=limite_vencido)
    ).count()
    cert_a_vencer = Arbitro.objects.filter(
        data_ultimo_exame__gte=limite_vencido,
        data_ultimo_exame__lte=limite_avencer_max,
    ).count()
    cert_regularizados = Arbitro.objects.filter(
        data_ultimo_exame__gt=limite_avencer_max,
    ).count()

    vencidos_q = Q(data_ultimo_exame__isnull=True) | Q(
        data_ultimo_exame__lt=limite_vencido
    )
    exames_pendentes = Arbitro.objects.filter(
        Q(estado_aptidao="aguardando_exame") | vencidos_q
    ).count()

    cert_sum = cert_vencidos + cert_a_vencer + cert_regularizados

    if total_arbitros:
        cert_pct_regular = int(round(100 * cert_regularizados / total_arbitros))
    else:
        cert_pct_regular = 0

    if cert_sum > 0:
        r = 360.0 * cert_regularizados / cert_sum
        v = 360.0 * cert_vencidos / cert_sum
        cert_conic_gradient = (
            f"conic-gradient(from -90deg, #0d59f2 0deg {r}deg, "
            f"#ef4444 {r}deg {r + v}deg, #cbd5e1 {r + v}deg 360deg)"
        )
    else:
        cert_conic_gradient = "conic-gradient(#e2e8f0 0deg 360deg)"

    proximas_marcacoes = (
        MarcacaoExame.objects.filter(data_hora_consulta__gte=now_dt)
        .select_related("arbitro", "prestador")
        .prefetch_related("itens")
        .order_by("data_hora_consulta")[:12]
    )
    proximos_exames = []
    for m in proximas_marcacoes:
        itens = list(m.itens.all())
        tipo = (
            ", ".join(i.exame_nome for i in itens[:3])
            if itens
            else "Consulta agendada"
        )
        if len(itens) > 3:
            tipo += "…"
        proximos_exames.append(
            {
                "quando": m.data_hora_consulta,
                "arbitro_nome": m.arbitro.nome_completo,
                "iniciais": _arbitro_iniciais(m.arbitro.nome_completo),
                "tipo_exame": tipo,
                "tipo_badge_class": _badge_class_exame(
                    itens[0].exame_nome if itens else tipo
                ),
                "prestador_nome": m.prestador.nome_completo
                if m.prestador_id
                else "—",
                "detalhe_url": reverse("exames:marcacao_detalhe", kwargs={"pk": m.pk}),
            }
        )

    alertas_raw = []
    for a in (
        Arbitro.objects.filter(
            data_ultimo_exame__gte=limite_vencido,
            data_ultimo_exame__lte=limite_avencer_max,
        )
        .only("nome_completo", "data_ultimo_exame")
        .order_by("-data_ultimo_exame")[:24]
    ):
        limite = a.data_ultimo_exame + timedelta(days=365)
        dias = (limite - today).days
        if 0 <= dias <= 30:
            alertas_raw.append(
                {
                    "nome": a.nome_completo,
                    "dias": dias,
                    "urgencia": "red" if dias <= 7 else "amber",
                }
            )
    alertas_raw.sort(key=lambda x: x["dias"])
    alertas_vencimento = alertas_raw[:6]

    context = {
        "total_arbitros": total_arbitros,
        "exames_pendentes": exames_pendentes,
        "exames_realizados_mes": exames_realizados_mes,
        "prestadores_ativos": prestadores_ativos,
        "epoca_atual": epoca_atual,
        "cert_regularizados": cert_regularizados,
        "cert_vencidos": cert_vencidos,
        "cert_a_vencer": cert_a_vencer,
        "cert_pct_regular": cert_pct_regular,
        "cert_conic_gradient": cert_conic_gradient,
        "proximos_exames": proximos_exames,
        "alertas_vencimento": alertas_vencimento,
    }
    return render(request, "core/dashboard.html", context)


@login_required
def profile_view(request):
    """Página de perfil do utilizador autenticado (com upload de avatar)."""
    user = request.user
    perfil, _ = PerfilConta.objects.get_or_create(
        user=user,
        defaults={"forcar_troca_password": False},
    )
    password_form = PasswordChangeForm(user=user)

    if request.method == "POST" and request.FILES.get("avatar"):
        avatar = request.FILES["avatar"]
        if avatar.size > 5 * 1024 * 1024:
            messages.error(request, "A imagem excede 5MB.")
            return redirect("core:profile")
        ctype = (getattr(avatar, "content_type", "") or "").lower()
        if not ctype.startswith("image/"):
            messages.error(request, "Envie um ficheiro de imagem válido (JPG, PNG, GIF, WEBP).")
            return redirect("core:profile")
        perfil.avatar = avatar
        perfil.save(update_fields=["avatar"])
        messages.success(request, "Foto de perfil atualizada com sucesso.")
        return redirect("core:profile")
    if request.method == "POST" and request.POST.get("action") == "change_password":
        password_form = PasswordChangeForm(user=user, data=request.POST)
        if password_form.is_valid():
            user = password_form.save()
            update_session_auth_hash(request, user)
            PerfilConta.objects.filter(user=user).update(forcar_troca_password=False)
            messages.success(request, "Palavra-passe atualizada com sucesso.")
            return redirect("core:profile")
        messages.error(request, "Não foi possível atualizar a palavra-passe. Verifique os campos.")

    for f in password_form.fields.values():
        css = (
            "w-full rounded-lg border border-slate-200 dark:border-slate-700 "
            "bg-white dark:bg-slate-900 px-4 py-2.5 text-sm text-slate-900 dark:text-slate-100 "
            "focus:ring-2 focus:ring-primary/50 focus:border-primary"
        )
        f.widget.attrs["class"] = css

    context = {
        "user": user,
        "perfil": perfil,
        "password_form": password_form,
    }
    return render(request, "core/profile.html", context)


class PasswordChangeComFlagView(PasswordChangeView):
    """
    Igual à vista predefinida do Django, mas limpa a obrigação de troca
    após alteração bem-sucedida (contas criadas pelo backoffice).
    """

    success_url = reverse_lazy("password_change_done")

    def form_valid(self, form):
        response = super().form_valid(form)
        PerfilConta.objects.filter(user=self.request.user).update(
            forcar_troca_password=False
        )
        return response


@login_required
@require_http_methods(["GET", "POST"])
def primeiro_acesso_password(request):
    """
    Troca obrigatória de palavra-passe no primeiro acesso.
    Página própria do frontend (não usa template do backend Django).
    """
    perfil, _ = PerfilConta.objects.get_or_create(
        user=request.user,
        defaults={"forcar_troca_password": False},
    )
    if not perfil.forcar_troca_password:
        return redirect("core:dashboard")

    if request.method == "POST":
        form = SetPasswordForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            perfil.forcar_troca_password = False
            perfil.save(update_fields=["forcar_troca_password"])
            messages.success(request, "Palavra-passe atualizada com sucesso.")
            return redirect("core:dashboard")
    else:
        form = SetPasswordForm(request.user)

    base_css = (
        "w-full h-12 px-4 rounded-xl border border-slate-200 dark:border-slate-700 "
        "bg-white dark:bg-slate-900 text-slate-900 dark:text-white "
        "focus:border-primary focus:ring-2 focus:ring-primary/20 outline-none transition-all "
        "placeholder:text-slate-400"
    )
    for field in form.fields.values():
        field.widget.attrs["class"] = base_css

    return render(request, "core/primeiro_acesso_password.html", {"form": form})


@login_required
@require_http_methods(["GET", "POST"])
def criar_utilizador(request):
    """
    Criação de utilizadores no backoffice. Restrito a superutilizadores.
    Palavra-passe gerada automaticamente, enviada por e-mail; primeira sessão
    obriga a definir uma nova palavra-passe.
    """
    if not request.user.is_superuser:
        raise PermissionDenied
    if request.method == "POST":
        form = CriarUtilizadorForm(request.POST)
        if form.is_valid():
            email = form.cleaned_data["email"]
            uname = username_a_partir_do_email(email)
            senha_temp = gerar_palavra_passe_segura(email)
            user = None
            try:
                with transaction.atomic():
                    user = User.objects.create_user(
                        username=uname,
                        email=email,
                        password=senha_temp,
                    )
                    user.first_name = (form.cleaned_data.get("first_name") or "").strip()
                    user.last_name = (form.cleaned_data.get("last_name") or "").strip()
                    user.is_active = bool(form.cleaned_data.get("is_active", True))
                    user.is_staff = bool(form.cleaned_data.get("is_staff", False))
                    user.is_superuser = bool(form.cleaned_data.get("is_superuser", False))
                    if user.is_superuser:
                        user.is_staff = True
                    user.save(
                        update_fields=[
                            "first_name",
                            "last_name",
                            "is_active",
                            "is_staff",
                            "is_superuser",
                        ]
                    )
                    PerfilConta.objects.create(
                        user=user,
                        forcar_troca_password=True,
                        departamento=(
                            form.cleaned_data["departamento"].nome
                            if form.cleaned_data.get("departamento")
                            else ""
                        ),
                    )
                    user.groups.set(form.cleaned_data.get("groups", []))
            except Exception as exc:
                if user is not None and user.pk:
                    user.delete()
                messages.error(
                    request,
                    "Não foi possível concluir o registo. "
                    f"Verifique a configuração de e-mail ou tente novamente. ({exc})",
                )
            else:
                login_url = absolute_url("/", request=request)
                corpo = (
                    f"Olá,\n\n"
                    f"Foi criada uma conta no Gol Health para o e-mail {email}.\n\n"
                    f"Aceda à plataforma em:\n{login_url}\n\n"
                    f"Palavra-passe temporária: {senha_temp}\n\n"
                    f"Utilize o e-mail acima para iniciar sessão. Por motivos de segurança, "
                    f"no primeiro acesso terá de definir uma nova palavra-passe antes de "
                    f"utilizar o sistema.\n\n"
                    f"— Gol Health"
                )
                try:
                    send_mail(
                        subject="Gol Health — acesso à sua conta",
                        message=corpo,
                        from_email=getattr(
                            settings,
                            "DEFAULT_FROM_EMAIL",
                            "webmaster@localhost",
                        ),
                        recipient_list=[email],
                        fail_silently=False,
                    )
                    messages.success(
                        request,
                        f"Utilizador criado. Foi enviada uma palavra-passe temporária para {email}.",
                    )
                except Exception as exc_mail:
                    messages.warning(
                        request,
                        "Utilizador criado, mas não foi possível enviar o e-mail agora. "
                        f"Use esta palavra-passe temporária: {senha_temp} "
                        f"(erro de e-mail: {exc_mail}).",
                    )
                return redirect("core:criar_utilizador")
    else:
        form = CriarUtilizadorForm()
    return render(
        request,
        "core/utilizador_novo.html",
        {"form": form},
    )


@login_required
def utilizadores_dashboard(request):
    """Dashboard/listagem de utilizadores da plataforma (apenas superutilizador)."""
    if not request.user.is_superuser:
        raise PermissionDenied
    q = (request.GET.get("q") or "").strip()
    users = (
        User.objects.all()
        .select_related("perfil_conta")
        .prefetch_related("groups")
        .annotate(
            departamento_nome=Coalesce(
                "perfil_conta__departamento",
                Value(""),
            )
        )
        .order_by("-date_joined", "-id")
    )
    if q:
        users = users.filter(
            Q(username__icontains=q)
            | Q(email__icontains=q)
            | Q(first_name__icontains=q)
            | Q(last_name__icontains=q)
            | Q(perfil_conta__departamento__icontains=q)
        )

    total_utilizadores = User.objects.count()
    total_ativos = User.objects.filter(is_active=True).count()
    total_inativos = total_utilizadores - total_ativos
    total_equipa = User.objects.filter(is_staff=True).count()
    total_superusers = User.objects.filter(is_superuser=True).count()
    total_sem_grupo = User.objects.filter(groups__isnull=True).distinct().count()
    total_troca_password_pendente = PerfilConta.objects.filter(
        forcar_troca_password=True
    ).count()
    total_com_avatar = PerfilConta.objects.exclude(avatar="").exclude(
        avatar__isnull=True
    ).count()

    departamentos_raw = (
        PerfilConta.objects.exclude(departamento__isnull=True)
        .exclude(departamento__exact="")
        .values("departamento")
        .annotate(total=Count("id"))
        .order_by("-total", "departamento")
    )
    departamentos = list(departamentos_raw[:6])
    dep_total_registos = sum(item["total"] for item in departamentos) or 1
    for item in departamentos:
        item["percent"] = int(round((item["total"] * 100) / dep_total_registos))

    utilizadores_recentes = (
        User.objects.select_related("perfil_conta")
        .prefetch_related("groups")
        .order_by("-date_joined", "-id")[:5]
    )

    return render(
        request,
        "core/utilizadores_dashboard.html",
        {
            "users": users,
            "search_q": q,
            "total_utilizadores": total_utilizadores,
            "total_ativos": total_ativos,
            "total_inativos": total_inativos,
            "total_equipa": total_equipa,
            "total_superusers": total_superusers,
            "total_sem_grupo": total_sem_grupo,
            "total_troca_password_pendente": total_troca_password_pendente,
            "total_com_avatar": total_com_avatar,
            "departamentos": departamentos,
            "utilizadores_recentes": utilizadores_recentes,
        },
    )


@login_required
@require_http_methods(["GET", "POST"])
def editar_utilizador(request, pk: int):
    """Editar dados e permissões de um utilizador (apenas superutilizador)."""
    if not request.user.is_superuser:
        raise PermissionDenied
    user_obj = get_object_or_404(User, pk=pk)
    perfil, _ = PerfilConta.objects.get_or_create(
        user=user_obj, defaults={"forcar_troca_password": False}
    )

    if request.method == "POST":
        form = EditarUtilizadorForm(request.POST, user_instance=user_obj)
        if form.is_valid():
            email = form.cleaned_data["email"]
            user_obj.email = email
            user_obj.username = username_a_partir_do_email(email)
            user_obj.first_name = (form.cleaned_data.get("first_name") or "").strip()
            user_obj.last_name = (form.cleaned_data.get("last_name") or "").strip()
            user_obj.is_active = bool(form.cleaned_data.get("is_active"))
            user_obj.is_staff = bool(form.cleaned_data.get("is_staff"))
            user_obj.is_superuser = bool(form.cleaned_data.get("is_superuser"))
            if user_obj.is_superuser:
                user_obj.is_staff = True
            user_obj.save()
            user_obj.groups.set(form.cleaned_data.get("groups", []))
            perfil.departamento = (
                form.cleaned_data["departamento"].nome
                if form.cleaned_data.get("departamento")
                else ""
            )
            perfil.save(update_fields=["departamento"])
            messages.success(request, "Utilizador atualizado com sucesso.")
            return redirect("core:utilizadores_dashboard")
    else:
        dep_obj = (
            Departamento.objects.filter(nome__iexact=(perfil.departamento or "")).first()
            if perfil.departamento
            else None
        )
        form = EditarUtilizadorForm(
            user_instance=user_obj,
            initial={
                "first_name": user_obj.first_name,
                "last_name": user_obj.last_name,
                "departamento": dep_obj.pk if dep_obj else None,
                "email": user_obj.email,
                "groups": list(user_obj.groups.values_list("pk", flat=True)),
                "is_active": user_obj.is_active,
                "is_staff": user_obj.is_staff,
                "is_superuser": user_obj.is_superuser,
            },
        )

    return render(
        request,
        "core/utilizador_editar.html",
        {"form": form, "user_obj": user_obj},
    )


@login_required
@require_POST
def utilizador_toggle_ativo(request, pk: int):
    """Ativa/desativa utilizador de forma rápida (apenas superutilizador)."""
    if not request.user.is_superuser:
        raise PermissionDenied
    user_obj = get_object_or_404(User, pk=pk)
    if user_obj.pk == request.user.pk:
        messages.error(request, "Não pode desativar a sua própria conta.")
        return redirect("core:utilizadores_dashboard")
    user_obj.is_active = not user_obj.is_active
    user_obj.save(update_fields=["is_active"])
    if user_obj.is_active:
        messages.success(request, f'Utilizador "{user_obj.email or user_obj.username}" ativado.')
    else:
        messages.success(request, f'Utilizador "{user_obj.email or user_obj.username}" desativado.')
    return redirect("core:utilizadores_dashboard")


@login_required
@require_POST
def utilizadores_apagar_em_massa(request):
    """Apaga vários utilizadores de uma vez (apenas superutilizador)."""
    if not request.user.is_superuser:
        raise PermissionDenied

    q = (request.POST.get("q") or "").strip()
    redirect_url = reverse("core:utilizadores_dashboard")
    if q:
        redirect_url = f"{redirect_url}?{urlencode({'q': q})}"

    raw_ids = request.POST.getlist("user_ids")
    if not raw_ids:
        messages.error(request, "Selecione pelo menos um utilizador para apagar.")
        return redirect(redirect_url)

    try:
        user_ids = {int(x) for x in raw_ids}
    except (TypeError, ValueError):
        messages.error(request, "Seleção inválida.")
        return redirect(redirect_url)

    user_ids.discard(request.user.pk)
    if not user_ids:
        messages.error(request, "Não pode apagar a sua própria conta.")
        return redirect(redirect_url)

    candidatos = list(User.objects.filter(pk__in=user_ids))
    if not candidatos:
        messages.error(request, "Nenhum utilizador válido selecionado.")
        return redirect(redirect_url)

    superusers_a_apagar = [u for u in candidatos if u.is_superuser]
    if superusers_a_apagar:
        restantes = (
            User.objects.filter(is_superuser=True)
            .exclude(pk__in=[u.pk for u in candidatos])
            .count()
        )
        if restantes == 0:
            messages.error(
                request,
                "Não é possível apagar todos os superutilizadores da plataforma.",
            )
            candidatos = [u for u in candidatos if not u.is_superuser]
            if not candidatos:
                return redirect(redirect_url)

    apagados = len(candidatos)
    User.objects.filter(pk__in=[u.pk for u in candidatos]).delete()
    messages.success(
        request,
        f"{apagados} utilizador{'es' if apagados != 1 else ''} apagado{'s' if apagados != 1 else ''} com sucesso.",
    )
    return redirect(redirect_url)


@login_required
def novo_departamento(request):
    """
    Página para registar um novo departamento no core.
    """
    if request.method == "POST":
        redirect_to = request.POST.get("redirect_to") or None
        form = DepartamentoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Departamento registado com sucesso.")
            return redirect(redirect_to or "core:criar_utilizador")
        else:
            nome_erro = form.errors.get("nome", [None])[0]
            messages.error(request, nome_erro or "Erro ao guardar departamento.")
            if redirect_to:
                sep = "&" if "?" in redirect_to else "?"
                return redirect(f"{redirect_to}{sep}departamento_modal=1")
    else:
        form = DepartamentoForm(initial={"ativo": True})
    return render(request, "core/novo_departamento.html", {"form": form})


@login_required
def novo_departamento_ajax(request):
    """
    Cria um departamento via AJAX (sem recarregar a página).
    Espera POST com `nome` e `ativo` opcional.
    """
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método não permitido."}, status=405)

    form = DepartamentoForm(request.POST)
    if not form.is_valid():
        nome_errs = form.errors.get("nome", [])
        return JsonResponse({"ok": False, "errors": {"nome": nome_errs}}, status=400)

    departamento = form.save()
    return JsonResponse({"ok": True, "id": departamento.id, "nome": departamento.nome})


@login_required
def novo_exame(request):
    """
    Permite registar um novo exame a partir do `core`.
    Reutiliza o `ExameForm` e o template existente em `exames`.
    """
    if request.method == "POST":
        form = ExameForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Exame registado com sucesso.")
            return redirect("exames:dashboard")
    else:
        form = ExameForm()

    return render(request, "exames/novo.html", {"form": form})


@login_required
def novo_exame_ajax(request):
    """
    Cria um novo ato médico via AJAX (sem recarregar a página).
    Espera POST com `nome` e `ativo` opcional.
    """
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método não permitido."}, status=405)

    form = ExameForm(request.POST)
    if not form.is_valid():
        nome_errs = form.errors.get("nome", [])
        ativo_errs = form.errors.get("ativo", [])
        return JsonResponse(
            {"ok": False, "errors": {"nome": nome_errs, "ativo": ativo_errs}},
            status=400,
        )

    exame = form.save()
    return JsonResponse({"ok": True, "id": exame.pk, "nome": exame.nome})


@login_required
def novo_categoria(request):
    """
    Permite registar novas categorias para aparecerem no dropdown de árbitros.
    """
    if request.method == "POST":
        redirect_to = request.POST.get("redirect_to") or None
        form = CategoriaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Categoria registada com sucesso.")
            return redirect(redirect_to or "arbitro:novo")
        else:
            # Caso a criação venha via modal, não queremos navegar para outra página.
            nome_erro = form.errors.get("nome", [None])[0]
            messages.error(request, nome_erro or "Erro ao guardar categoria.")
            if redirect_to:
                sep = "&" if "?" in redirect_to else "?"
                return redirect(f"{redirect_to}{sep}categoria_modal=1")

    form = CategoriaForm()

    return render(request, "core/nova_categoria.html", {"form": form})


@login_required
def novo_modalidade(request):
    """
    Permite registar novas modalidades para aparecerem no dropdown de árbitros.
    """
    if request.method == "POST":
        redirect_to = request.POST.get("redirect_to") or None
        form = ModalidadeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Modalidade registada com sucesso.")
            return redirect(redirect_to or "arbitro:novo")
        else:
            nome_erro = form.errors.get("nome", [None])[0]
            messages.error(request, nome_erro or "Erro ao guardar modalidade.")
            if redirect_to:
                sep = "&" if "?" in redirect_to else "?"
                return redirect(f"{redirect_to}{sep}modalidade_modal=1")

    form = ModalidadeForm()
    return render(request, "core/nova_modalidade.html", {"form": form})


@login_required
def novo_categoria_ajax(request):
    """
    Cria uma categoria via AJAX (sem recarregar a página).
    Espera POST com `nome`.
    """
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método não permitido."}, status=405)

    form = CategoriaForm(request.POST)
    if not form.is_valid():
        nome_errs = form.errors.get("nome", [])
        return JsonResponse({"ok": False, "errors": {"nome": nome_errs}}, status=400)

    categoria = form.save()
    return JsonResponse({"ok": True, "id": categoria.id, "nome": categoria.nome})


@login_required
def novo_modalidade_ajax(request):
    """
    Cria uma modalidade via AJAX (sem recarregar a página).
    Espera POST com `nome`.
    """
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Método não permitido."}, status=405)

    form = ModalidadeForm(request.POST)
    if not form.is_valid():
        nome_errs = form.errors.get("nome", [])
        return JsonResponse({"ok": False, "errors": {"nome": nome_errs}}, status=400)

    modalidade = form.save()
    return JsonResponse({"ok": True, "id": modalidade.id, "nome": modalidade.nome})


@login_required
def novo_epoca(request):
    """
    Página para registar uma nova época no `core`.
    """
    if request.method == "POST":
        form = EpocaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Época registada com sucesso.")
            return redirect("exames:dashboard")
    else:
        form = EpocaForm()

    return render(request, "core/nova_epoca.html", {"form": form})


@login_required
@require_POST
def novo_epoca_ajax(request):
    """
    Cria uma época via AJAX (sem navegar para `/epocas/novo/`).
    Espera POST com os campos do `EpocaForm`.
    """
    form = EpocaForm(request.POST)
    if not form.is_valid():
        errors = {
            "nome": form.errors.get("nome", []),
            "inicio_data": form.errors.get("inicio_data", []),
            "fim_data": form.errors.get("fim_data", []),
            "ativo": form.errors.get("ativo", []),
        }
        return JsonResponse({"ok": False, "errors": errors}, status=400)

    epoca = form.save()
    return JsonResponse({"ok": True, "id": epoca.pk, "nome": epoca.nome})


@login_required
def epoca_detalhe(request, pk: int):
    """
    Mostra detalhes da época e permite mudar o estado ativa/inativa.
    """
    epoca = get_object_or_404(Epoca, pk=pk)

    if request.method == "POST":
        form = EpocaAtivoForm(request.POST, instance=epoca)
        if form.is_valid():
            form.save()  # garante desativação das outras quando ativo=True
            messages.success(request, "Estado da época atualizado com sucesso.")
            return redirect("core:epoca_detalhe", pk=epoca.pk)
    else:
        form = EpocaAtivoForm(instance=epoca)

    epoca_associacao_ids = EpocaAssociacao.objects.filter(epoca=epoca).values_list(
        "associacao_id", flat=True
    )

    associacoes = (
        Associacao.objects.filter(id__in=epoca_associacao_ids)
        .annotate(exames_count=Count("exames_valores"))
        .order_by("-criado_em")
    )

    associacoes_disponiveis = (
        Associacao.objects.exclude(id__in=epoca_associacao_ids)
        .order_by("nome_completo")
    )

    arbitros = (
        Arbitro.objects.filter(epocas__epoca=epoca, ativo=True)
        .select_related("associacao_futebol", "categoria", "modalidade")
        .order_by("-criado_em")
    )

    return render(
        request,
        "core/epoca_detalhe.html",
        {
            "epoca": epoca,
            "form": form,
            "associacoes": associacoes,
            "associacoes_disponiveis": associacoes_disponiveis,
            "arbitros": arbitros,
        },
    )


@login_required
@require_POST
def epoca_associacao_add_ajax(request, pk: int):
    """
    Associa uma `Associacao` à `Epoca`.
    """
    epoca = get_object_or_404(Epoca, pk=pk)
    associacao_id = request.POST.get("associacao_id")

    if not associacao_id:
        return JsonResponse({"ok": False, "error": "Selecione uma associação."}, status=400)

    associacao = get_object_or_404(Associacao, pk=associacao_id)

    obj, created = EpocaAssociacao.objects.get_or_create(epoca=epoca, associacao=associacao)
    return JsonResponse({"ok": True, "created": created, "id": obj.pk})


def _normalize_header(value: str) -> str:
    """
    Normaliza cabeçalhos do Excel:
    - remove acentos
    - coloca em minúsculas
    - remove espaços extra
    """
    value = str(value or "").strip().lower()
    value = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    value = " ".join(value.split())
    return value


def _parse_date_maybe(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        raw = value.strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue
    return None


@login_required
@require_http_methods(["GET", "POST"])
def epoca_associacao_import_arbitros(request, epoca_pk: int, associacao_pk: int):
    epoca = get_object_or_404(Epoca, pk=epoca_pk)
    associacao = get_object_or_404(Associacao, pk=associacao_pk)

    # Garante que a associação está realmente ligada à época.
    if not EpocaAssociacao.objects.filter(epoca=epoca, associacao=associacao).exists():
        messages.error(request, "Esta associação não está ligada a esta época.")
        return redirect("core:epoca_detalhe", pk=epoca.pk)

    arbitros_da_epoca = (
        Arbitro.objects.filter(epocas__epoca=epoca, associacao_futebol=associacao)
        .select_related("categoria", "modalidade", "associacao_futebol")
        .order_by("nome_completo")
    )

    if request.method == "POST":
        form = ImportArbitrosExcelForm(request.POST, request.FILES)
        if not form.is_valid():
            messages.error(request, "Ficheiro inválido. Tente novamente.")
            return render(
                request,
                "core/importar_arbitros_excel.html",
                {"epoca": epoca, "associacao": associacao, "form": form},
            )

        uploaded = form.cleaned_data["arquivo"]
        raw_bytes = uploaded.read()
        wb = openpyxl.load_workbook(BytesIO(raw_bytes), data_only=True)
        ws = wb.active

        # Cabeçalhos esperados (variações aceites por normalização).
        headers = [cell.value for cell in ws[1]]
        header_map = {}
        for idx, h in enumerate(headers):
            norm = _normalize_header(h)
            if norm:
                header_map[norm] = idx

        def get_by_keys(row, keys):
            for k in keys:
                nk = _normalize_header(k)
                if nk in header_map:
                    return row[header_map[nk]]
            return None

        inserted = 0
        updated = 0
        skipped = 0
        errors = []

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=2, values_only=True), start=2
        ):
            nome_completo = get_by_keys(
                row, ["nome completo", "nome_completo", "nome", "nome arbitro"]
            )
            # Fallback quando o Excel não tiver cabeçalhos compatíveis:
            # assume a primeira coluna como nome do árbitro.
            if not nome_completo and row and len(row) > 0:
                nome_completo = row[0]

            if not nome_completo:
                skipped += 1
                continue

            nome_completo = str(nome_completo).strip()
            data_nascimento = _parse_date_maybe(
                get_by_keys(
                    row,
                    [
                        "data nascimento",
                        "data de nascimento",
                        "data_nascimento",
                        "nascimento",
                    ],
                )
            )
            if data_nascimento is None and row and len(row) > 1:
                data_nascimento = _parse_date_maybe(row[1])

            cpf_nif_raw = get_by_keys(row, ["cpf nif", "cpf_nif", "nif", "cpf"])
            cpf_nif = str(cpf_nif_raw).strip() if cpf_nif_raw not in (None, "") else ""
            cpf_nif = cpf_nif.replace(" ", "").replace(".", "").replace("-", "")
            if not cpf_nif and row and len(row) > 2 and row[2] not in (None, ""):
                cpf_nif = str(row[2]).strip().replace(" ", "").replace(".", "").replace("-", "")

            sexo_raw = get_by_keys(row, ["sexo"])
            sexo_raw_norm = (str(sexo_raw).strip().lower() if sexo_raw else "")
            sexo = ""
            if sexo_raw_norm:
                if sexo_raw_norm.startswith("m"):
                    sexo = "M"
                elif sexo_raw_norm.startswith("f"):
                    sexo = "F"
                elif "masc" in sexo_raw_norm:
                    sexo = "M"
                elif "fem" in sexo_raw_norm:
                    sexo = "F"
            if not sexo and row and len(row) > 3 and row[3] not in (None, ""):
                sexo_raw_norm = str(row[3]).strip().lower()
                if sexo_raw_norm.startswith("m"):
                    sexo = "M"
                elif sexo_raw_norm.startswith("f"):
                    sexo = "F"
                elif "masc" in sexo_raw_norm:
                    sexo = "M"
                elif "fem" in sexo_raw_norm:
                    sexo = "F"

            telefone = get_by_keys(row, ["telemovel", "telemóvel", "telefone", "tel"])
            email = get_by_keys(row, ["email", "e-mail"])
            codigo_postal = get_by_keys(row, ["codigo postal", "código postal", "codigo_postal", "cp"])
            localidade = get_by_keys(row, ["localidade"])
            rua_avenida = get_by_keys(row, ["rua avenida", "rua/avenida", "rua / avenida", "rua_avenida"])
            numero = get_by_keys(row, ["numero", "nº"])
            complemento = get_by_keys(row, ["complemento"])
            observacoes_morada = get_by_keys(row, ["observacoes", "observacoes morada", "observações", "observacoes_morada"])
            cidade = get_by_keys(row, ["cidade"])

            categoria_name = get_by_keys(row, ["categoria"])
            modalidade_name = get_by_keys(row, ["modalidade"])

            # Data do último exame (opcional)
            data_ultimo_exame = _parse_date_maybe(
                get_by_keys(row, ["data ultimo exame", "data do ultimo exame", "data_ultimo_exame"])
            )

            # Garante Categoria/Modalidade existentes (cria se necessário).
            categoria_obj = None
            if categoria_name:
                categoria_obj, _ = Categoria.objects.get_or_create(nome=str(categoria_name).strip())

            modalidade_obj = None
            if modalidade_name:
                modalidade_obj, _ = Modalidade.objects.get_or_create(nome=str(modalidade_name).strip())

            # Deduplicação global (por CPF/NIF se existir; senão por nome+data, e por último apenas nome).
            qs = Arbitro.objects.all()
            if cpf_nif:
                qs = qs.filter(cpf_nif=cpf_nif)
            elif nome_completo and data_nascimento:
                qs = qs.filter(nome_completo=nome_completo, data_nascimento=data_nascimento)
            else:
                qs = qs.filter(nome_completo=nome_completo)

            arbitro = qs.order_by("pk").first()
            try:
                if arbitro:
                    updates = {}

                    # Nome é sempre atualizado (vindo do Excel).
                    if arbitro.nome_completo != nome_completo:
                        updates["nome_completo"] = nome_completo

                    # Só atualiza campos opcionais se o Excel trouxer valor.
                    if data_nascimento is not None and arbitro.data_nascimento != data_nascimento:
                        updates["data_nascimento"] = data_nascimento

                    if cpf_nif:
                        if arbitro.cpf_nif != cpf_nif:
                            updates["cpf_nif"] = cpf_nif

                    if sexo:
                        if arbitro.sexo != sexo:
                            updates["sexo"] = sexo

                    if email:
                        if arbitro.email != email:
                            updates["email"] = email

                    if telefone:
                        if arbitro.telefone != telefone:
                            updates["telefone"] = telefone

                    if codigo_postal:
                        if arbitro.codigo_postal != codigo_postal:
                            updates["codigo_postal"] = codigo_postal

                    if localidade:
                        if arbitro.localidade != localidade:
                            updates["localidade"] = localidade

                    if rua_avenida:
                        if arbitro.rua_avenida != rua_avenida:
                            updates["rua_avenida"] = rua_avenida

                    if numero not in (None, ""):
                        numero_str = str(numero).strip()
                        if arbitro.numero != numero_str:
                            updates["numero"] = numero_str

                    if complemento:
                        if arbitro.complemento != complemento:
                            updates["complemento"] = complemento

                    if observacoes_morada:
                        if arbitro.observacoes_morada != observacoes_morada:
                            updates["observacoes_morada"] = observacoes_morada

                    if cidade:
                        if arbitro.cidade != cidade:
                            updates["cidade"] = cidade

                    if categoria_obj is not None:
                        if arbitro.categoria_id != categoria_obj.id:
                            updates["categoria"] = categoria_obj

                    if modalidade_obj is not None:
                        if arbitro.modalidade_id != modalidade_obj.id:
                            updates["modalidade"] = modalidade_obj

                    if data_ultimo_exame is not None:
                        if arbitro.data_ultimo_exame != data_ultimo_exame:
                            updates["data_ultimo_exame"] = data_ultimo_exame

                    # Reatribui associação à época/associação importada.
                    if arbitro.associacao_futebol_id != associacao.id:
                        updates["associacao_futebol"] = associacao

                    if updates:
                        for k, v in updates.items():
                            setattr(arbitro, k, v)
                        arbitro.save()
                        updated += 1
                else:
                    arbitro = Arbitro.objects.create(
                        nome_completo=nome_completo,
                        data_nascimento=data_nascimento,
                        cpf_nif=cpf_nif,
                        sexo=sexo,
                        email=email or "",
                        telefone=telefone or "",
                        codigo_postal=codigo_postal or "",
                        localidade=localidade or "",
                        rua_avenida=rua_avenida or "",
                        numero=str(numero).strip() if numero not in (None, "") else "",
                        complemento=complemento or "",
                        observacoes_morada=observacoes_morada or "",
                        cidade=cidade or "",
                        categoria=categoria_obj,
                        associacao_futebol=associacao,
                        modalidade=modalidade_obj,
                        data_ultimo_exame=data_ultimo_exame,
                    )
                    inserted += 1

                EpocaArbitro.objects.get_or_create(epoca=epoca, arbitro=arbitro)
            except Exception as exc:
                errors.append(f"Linha {row_idx}: {exc}")

        # Recalcula lista após importação.
        arbitros_da_epoca = (
            Arbitro.objects.filter(epocas__epoca=epoca, associacao_futebol=associacao)
            .select_related("categoria", "modalidade", "associacao_futebol")
            .order_by("nome_completo")
        )

        result = {
            "inserted": inserted,
            "updated": updated,
            "skipped": skipped,
            "error_count": len(errors),
            "errors": errors[:20],
        }

        context = {
            "epoca": epoca,
            "associacao": associacao,
            "form": ImportArbitrosExcelForm(),
            "result": result,
            "arbitros": arbitros_da_epoca,
        }
        return render(request, "core/importar_arbitros_excel.html", context)

    form = ImportArbitrosExcelForm()
    return render(
        request,
        "core/importar_arbitros_excel.html",
        {"epoca": epoca, "associacao": associacao, "form": form, "arbitros": arbitros_da_epoca},
    )


@login_required
def epoca_associacao_arbitros(request, epoca_pk: int, associacao_pk: int):
    """
    Lista árbitros ligados à combinação (época + associação),
    com botão para abrir a importação.
    """
    epoca = get_object_or_404(Epoca, pk=epoca_pk)
    associacao = get_object_or_404(Associacao, pk=associacao_pk)

    # Garante que a associação está realmente ligada à época.
    if not EpocaAssociacao.objects.filter(epoca=epoca, associacao=associacao).exists():
        messages.error(request, "Esta associação não está ligada a esta época.")
        return redirect("core:epoca_detalhe", pk=epoca.pk)

    arbitros = (
        Arbitro.objects.filter(epocas__epoca=epoca, associacao_futebol=associacao, ativo=True)
        .select_related("categoria", "modalidade", "associacao_futebol")
        .order_by("nome_completo")
    )

    return render(
        request,
        "core/epoca_associacao_arbitros_dashboard.html",
        {
            "epoca": epoca,
            "associacao": associacao,
            "arbitros": arbitros,
        },
    )
