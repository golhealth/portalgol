from __future__ import annotations

import unicodedata
from io import BytesIO

import openpyxl
from django.db import transaction

from .models import Exame


def _normalize_header(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = unicodedata.normalize("NFKD", text)
    return "".join(ch for ch in text if not unicodedata.combining(ch))


def _cell_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _parse_codigo(value):
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value if value > 0 else None
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        return int(value) if value > 0 else None
    raw = _cell_str(value).replace(",", ".")
    if not raw:
        return None
    try:
        if "." in raw:
            return int(float(raw))
        return int(raw)
    except (TypeError, ValueError):
        return None


def _find_header_row(ws, *, max_scan_rows: int = 15) -> tuple[int, dict[str, int]] | tuple[None, None]:
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True),
        start=1,
    ):
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            norm = _normalize_header(cell)
            if norm in ("codigo", "cod"):
                col_map["codigo"] = col_idx
            elif norm in ("descritivo", "descricao", "nome", "nome do ato medico"):
                col_map["descritivo"] = col_idx
            elif norm in ("abreviatura", "abrev", "sigla"):
                col_map["abreviatura"] = col_idx
        if "codigo" in col_map and "descritivo" in col_map:
            if "abreviatura" not in col_map:
                col_map["abreviatura"] = col_map["descritivo"] + 1
            return row_idx, col_map
    return None, None


def importar_atos_excel(raw_bytes: bytes) -> dict:
    """
    Importa atos médicos a partir de .xlsx.
    Usa apenas as colunas Código, Descritivo e Abreviatura (ignora Observações).
    """
    wb = openpyxl.load_workbook(BytesIO(raw_bytes), data_only=True)
    ws = wb.active

    header_row, col_map = _find_header_row(ws)
    if not header_row or not col_map:
        wb.close()
        raise ValueError(
            "Não foi possível encontrar o cabeçalho com as colunas "
            "Código, Descritivo e Abreviatura."
        )

    inserted = 0
    updated = 0
    skipped = 0
    errors: list[str] = []

    with transaction.atomic():
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if not row:
                skipped += 1
                continue

            codigo = _parse_codigo(
                row[col_map["codigo"]] if col_map["codigo"] < len(row) else None
            )
            descritivo = _cell_str(
                row[col_map["descritivo"]] if col_map["descritivo"] < len(row) else None
            )
            abreviatura = _cell_str(
                row[col_map["abreviatura"]] if col_map["abreviatura"] < len(row) else None
            )

            if codigo is None and not descritivo:
                skipped += 1
                continue

            if not descritivo:
                errors.append(f"Linha {row_idx}: descritivo em falta (código {codigo}).")
                continue

            if codigo is None:
                errors.append(f"Linha {row_idx}: código em falta para «{descritivo}».")
                continue

            defaults = {
                "nome": descritivo,
                "abreviatura": abreviatura[:50],
                "ativo": True,
            }

            existing = Exame.objects.filter(codigo=codigo).first()
            if existing:
                changed = False
                for field, value in defaults.items():
                    if getattr(existing, field) != value:
                        setattr(existing, field, value)
                        changed = True
                if changed:
                    existing.save()
                    updated += 1
                else:
                    skipped += 1
                continue

            conflict = Exame.objects.filter(nome__iexact=descritivo).exclude(codigo=codigo).first()
            if conflict:
                conflict.codigo = codigo
                conflict.abreviatura = abreviatura[:50]
                conflict.nome = descritivo
                conflict.ativo = True
                conflict.save()
                updated += 1
                continue

            Exame.objects.create(codigo=codigo, **defaults)
            inserted += 1

    wb.close()
    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
    }
